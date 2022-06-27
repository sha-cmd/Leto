# -*- coding: utf-8 -*-
"""
Created on Thu Feb 27 11:51:15 2020
ATTENTION DE CHANGER exticket database quand le nom de la base de données
change
@author: rboyrie
"""
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from machine import do as doing
from machine import transcription as trn
from machine import request as rq

def make(month_start,month_end,year_start,year_end):
    """
    Fonction regroupant la création des valeurs de la performances
    dans le détail pour tous les process, ainsi que pour la vision globale
    du service informatique, c’est-à-dire de l’ensemble des tickets.
    """
    do(month_start,month_end,year_start,year_end,process=0)
    do(month_start,month_end,year_start,year_end,process=1)
    do(month_start,month_end,year_start,year_end,process=2)
    do(month_start,month_end,year_start,year_end,process=3)
    do(month_start,month_end,year_start,year_end,process=4)


def do(month_start, month_ended, year_start, year_end, mode="replace"\
       , process=1):
    """
    Calcul les performances pour une périodes données, même trans-annuelle.
    Retourne des valeurs pour les performances globales mais aussi par type 
    ressource ou support, et également pour chaque process (pour certaines 
    valeurs (regarder le code pour plus d’information). Car en entrée peut-être
    donné le numéro de process, ce qui sert à observer les supports et les 
    ressources à l’intérieur d’un process donné en entrée. Le tout est écrit
    en base de données et dans un fichier excel portant un suffixe
    """
    mois = ['Janvier','Février','Mars','Avril','Mai','Juin',
                               'Juillet','Août','Septembre',
                               'Octobre','Novembre'
                               ,'Décembre']
    da = pd.DataFrame()
    for year in range (year_start, year_end+1):
        if (year < year_end):
            month_end = 12
        else:
            month_end = month_ended
        for i in range (month_start,month_end+1):
            data = rq.data_bymonth_byprocess(i, year, process)

            ## Mixage des données de résolutions Tr et Trc pour tous types
            ## de groupes : Global, Incidents(Supports), Demandes(Ressources)
            ## Et puis tous les process.
            data_data = data['Tr'].append(data['Trc'], ignore_index=True)
            data_i_mixed = data.loc[(data['Process'].str.contains('i rc ' \
                                     + str(process)))]['Tr']\
        .append(data.loc[(data['Process'].str.contains('i c '\
                          + str(process)))]['Trc'], ignore_index=True)
            data_d_mixed = data.loc[(data['Process'].str.contains('d rc ' \
                                     + str(process)))]['Tr']\
        .append(data.loc[(data['Process'].str.contains('d c '\
                          + str(process)))]['Trc'], ignore_index=True)
            data_1_mixed = data.loc[(data['Process']\
                                     .str.contains(' rc 1'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains(' c 1'))]['Trc'], \
                         ignore_index=True)# une espace avant " c 1" a ne 
        # surtout pas enlever car les tickets résolue et clôturé Trc ont bien
        # une clôture  et ne doivent pas être confondus avec les tickets qui 
        # sont résolus et clôturés en deux temps nommées rc dans la colonne
        # Process
            data_2_mixed = data.loc[(data['Process']\
                                     .str.contains(' rc 2'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains(' c 2'))]['Trc'],\
                         ignore_index=True)
            data_3_mixed = data.loc[(data['Process']\
                                     .str.contains(' rc 3'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains(' c 3'))]['Trc'], \
                         ignore_index=True)
            data_4_mixed = data.loc[(data['Process']\
                                     .str.contains(' rc 4'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains(' c 4'))]['Trc'], \
                         ignore_index=True)
            data_1_mixed_c = data.loc[(data['Process']\
                                       .str.contains(' rc 1'))]['Tc']
            data_2_mixed_c = data.loc[(data['Process']\
                                       .str.contains(' rc 2'))]['Tc']
            data_3_mixed_c = data.loc[(data['Process']\
                                       .str.contains(' rc 3'))]['Tc']
            data_4_mixed_c = data.loc[(data['Process']\
                                       .str.contains(' rc 4'))]['Tc']
            # Le process 0 est le groupe de l’ensemble de tous les tickets,
            # le global. Cela est utile dans le calcul à vision de la totalité.
            if process == 0 :
                data_i_mixed = data.loc[(data['Process']\
                                         .str.contains('i rc'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains('i c'))]['Trc'], \
                         ignore_index=True)
                data_d_mixed = data.loc[(data['Process']\
                                         .str.contains('d rc'))]['Tr']\
        .append(data.loc[(data['Process'].str.contains('d c'))]['Trc'], \
                         ignore_index=True)
            # Prise en charge par le Front Office Quantité, In SLA, Out SLA
            pecplfoQ = data.loc[data['clos']==1]['demande'].count() 
            pecplfoI = "%.4f" % (float(data.loc[(data['T1']\
                                                 .div(3600000000000)<=0.5)\
    &(data['clos']==1)]['demande'].count())\
                        /float(pecplfoQ)) if (pecplfoQ != 0) else 0
            pecplfoO = "%.4f" % (float(data.loc[(data['T1']\
                                                 .div(3600000000000)>0.5)\
        &(data['clos']==1)]['demande'].count())\
                                 /float(pecplfoQ)) if (pecplfoQ != 0) else 0
            # Question des Urgences, basse (quantité, in SLA, out SLA) et puis 
            # les urgences moyennes, hautes puis très hautes, avec à chaque
            # fois les questions de quantités et d’In SlA puis Out SLA
            ubq = data.loc[(data['clos']==1)&(data['urgence']==1)]['demande']\
            .count()
            ubq = float(ubq)
            ubi = "%.4f" % (float(data.loc[(data['Tt'].div(3600000000000)<=40)\
                                           &(data['clos']==1)\
                                           &(data['urgence']==1)]['demande']\
            .count())\
                    / float(ubq)) if (ubq != 0) else 0
            ubi = float(ubi)
            ubo = "%.4f" % (float(data.loc[(data['Tt'].div(3600000000000)>40)\
                                           &(data['clos']==1)\
                                           &(data['urgence']==1)]['demande']\
            .count())\
                    / float(ubq)) if (ubq != 0) else 0
            ubo = float(ubo)
            umq = data.loc[(data['clos']==1)&(data['urgence']==2)]['demande']\
            .count()
            umq = float(umq)
            umi = "%.4f" % (float(data.loc[(data['Tt'].div(3600000000000)<=16)\
                                           &(data['clos']==1)\
                                           &(data['urgence']==2)]['demande']\
            .count())\
                    / float(umq)) if (umq != 0) else 0
            umi = float(umi)
            umo = "%.4f" % (float(data.loc[(data['Tt']\
                                            .div(3600000000000)>16)\
        &(data['clos']==1)&(data['urgence']==2)]['demande'].count())\
                    / float(umq)) if (umq != 0) else 0
            umo = float(umo)
            uhq = data.loc[(data['clos']==1)&(data['urgence']==3)]['demande']\
            .count()
            uhq = float(uhq)
            uhi = "%.4f" % (float(data.loc[(data['Tt']\
                                            .div(3600000000000)<=2)\
        &(data['clos']==1)&(data['urgence']==3)]['demande'].count())\
                    / float(uhq)) if (uhq != 0) else 0
            uhi = float(uhi)
            uho = "%.4f" % (float(data.loc[(data['Tt']\
                                            .div(3600000000000)>2)\
        &(data['clos']==1)&(data['urgence']==3)]['demande'].count())\
                    / float(uhq)) if (uhq != 0) else 0
            uho = float(uho)
            utq = data.loc[(data['clos']==1)\
                           &(data['urgence']==4)]['demande'].count()
            utq = float(utq)
            uti = "%.4f" % (float(data.loc[(data['Tt']\
                                            .div(3600000000000)<=1)\
        &(data['clos']==1)&(data['urgence']==4)]['demande'].count())\
                    / float(utq)) if (utq != 0) else 0
            uti = float(uti)
            uto = "%.4f" % (float(data.loc[(data['Tt']\
                                            .div(3600000000000)>1)\
        &(data['clos']==1)&(data['urgence']==4)]['demande'].count())\
                    / float(utq)) if (utq != 0) else 0
            uto = float(uto) 
                            
            pecsq = data.loc[(data['clos']==1)&(data['Process']\
                             .str.contains('i'))]['demande'].count()
            pecsi = "%.4f" % (float(data.loc[(data['T1']<=(0.5*3600000000000))\
                                             &(data['clos']==1)\
                                             &(data['Process']\
                                               .str.contains('i'))]['T1']\
                                               .count())/float(pecsq)) \
                                               if (pecsq != 0) else 0
            pecso = "%.4f" % (float(data.loc[(data['T1']>(0.5*3600000000000))\
                                             &(data['clos']==1)\
                                             &(data['Process']\
                                               .str.contains('i'))]['T1']\
                                               .count())/float(pecsq)) \
                                               if (pecsq != 0) else 0
            pecdq = data.loc[(data['clos']==1)&(data['Process']\
                             .str.contains('d'))]['demande'].count()
            pecdi = "%.4f" % \
            (float(data.loc[(data['T1']<=(0.5*3600000000000))\
                            &(data['clos']==1)&(data['Process']\
                             .str.contains('d'))]['T1'].count())\
                                /float(pecdq)) if (pecdq != 0) else 0
            pecdo = "%.4f" % \
            (float(data.loc[(data['T1']>(0.5*3600000000000))&(data['clos']==1)\
                            &(data['Process']\
                              .str.contains('d'))]['T1'].count())\
                                /float(pecdq)) if (pecdq != 0) else 0
            traiq = data.loc[data['clos']==1]['Tt'].count()
            # Calcul historique des SLA et traitement qui ne tenait pas 
            # des urgences mais prenait comme seuil pour tout traitement
            # 8 heures. Cela est dû à l’absence de vision des process, qui 
            # était venue plus tard.
            #traii = "%.4f" % (float(data.loc[(data['Tt']<=(8*3600000000000))\
            # & (data['clos']==1)]['Tt'].count())/float(traiq))\
            # if (traiq != 0) else 0
            #traio = "%.4f" % (float(data.loc[(data['Tt']>(8*3600000000000))\
            # &(data['clos']==1)]['Tt'].count())/float(traiq))\
            # if (traiq != 0) else 0
            traii = "%.4f" % \
            (float(((ubq*ubi)+(umq*umi)+(uhq*uhi)+(utq*uti))) / float(traiq)) \
            if (traiq != 0) else 0
            traio = "%.4f" % \
            (float(((ubq*ubo)+(umq*umo)+(uhq*uho)+(utq*uto))) / float(traiq)) \
            if (traiq != 0) else 0
            ## KPI SLA SELON LA DEMANDE DU RESPONSABLE DE LA QUALITÉ DSI                     
            nbinc = data.loc[(data['Process']\
                              .str.contains('i'))]['demande'].count()
            nbdem = data.loc[(data['Process']\
                              .str.contains('d'))]['demande'].count()
            nbio1 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=1, nature='i')
            nbio2 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=2, nature='i')
            nbio3 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=3, nature='i')
            nbio4 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=4, nature='i')
            nbdo1 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=1, nature='d')
            nbdo2 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=2, nature='d')
            nbdo3 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=3, nature='d')
            nbdo4 = rq.opened_ticket_number_in_a_month(month=i, year=year, \
                                                       process=4, nature='d')
             # Cet indicateur est fonction du process
            dmapsi = "%.2f" %  (data.loc[(data['Process']\
                                          .str.contains(' i '))]['T1']\
                                            .div(3600000000000).mean())
            dmaci = "%.2f" % (pd.Series(data_i_mixed).div(3600000000000).mean() 
                            + data.loc[(data['Process']\
                                        .str.contains(' i '))]['T1']\
                                                .div(3600000000000).mean() 
                            + data.loc[(data['Process']\
                                        .str.contains(' i '))]['T2']\
                                        .div(3600000000000).mean())
            dmapsd = "%.2f" % (data.loc[(data['Process']\
                                         .str.contains(' d '))]['T1']\
                                        .div(3600000000000).mean())
            dmacd = "%.2f" % (pd.Series(data_d_mixed)\
                              .div(3600000000000).mean()  
                            + data.loc[(data['Process']\
                                        .str.contains(' d '))]['T1']\
                                        .div(3600000000000).mean()
                            + data.loc[(data['Process']\
                                        .str.contains(' d '))]['T2']\
                                        .div(3600000000000).mean())
            tri = "%.2f" % (float(cote_z(data_i_mixed))) 
            if (tri == 'Null') :
                tri = 0
            trd = "%.2f" % (float(cote_z(data_d_mixed))) 
            if (trd == 'Null') :
                trd = 0
            tr  = "%.2f" % (float(cote_z(data_data)))
            tr1 = "%.2f" % (float(cote_z(data_1_mixed)))
            tr2 = "%.2f" % (float(cote_z(data_2_mixed)))
            tr3 = "%.2f" % (float(cote_z(data_3_mixed)))
            tr4 = "%.2f" % (float(cote_z(data_4_mixed)))
            tc1 = "%.2f" % (float(cote_z(data_1_mixed_c)))
            tc2 = "%.2f" % (float(cote_z(data_2_mixed_c)))
            tc3 = "%.2f" % (float(cote_z(data_3_mixed_c)))
            tc4 = "%.2f" % (float(cote_z(data_4_mixed_c)))
            try:
                la_date = \
                pd.to_datetime(data['date_de_cloture']\
                               .iloc[0],format="%Y-%m-%d %H:%M:%S")
                la_date = pd.to_datetime('{}/{}'\
                                         .format(la_date.month,la_date.year)\
                                         ,format\
                                     ='%m/%Y').date()
            except: # Résoud les cas ou aucun ticket n’a aucune clôture i étant le mois
                la_date = pd.to_datetime('{}/{}'.format(i,year),format\
                                     ='%m/%Y').date()
            
            # Ajout des données calculées dans le cadre de données
            da = da.append({'mois': mois[i-1],
                            'date':la_date,
                   'Prise en charge (par le Front Office) Quantité':pecplfoQ,
                   'Prise en charge (par le Front Office) inside SLA':pecplfoI,
               'Prise en charge (par le Front Office) outside SLA':pecplfoO,
                       'SLA Urgence basse (traitement) Quantité':ubq,
                       'SLA Urgence basse (traitement) inside SLA':ubi,
                       'SLA Urgence basse (traitement) outside SLA':ubo,
                       'SLA Urgence moyenne (traitement) Quantité':umq,
                       'SLA Urgence moyenne (traitement) inside SLA':umi,
                       'SLA Urgence moyenne (traitement) outside SLA':umo,
                       'SLA Urgence Haute (traitement) Quantité':uhq,
                       'SLA Urgence Haute (traitement) inside SLA':uhi,
                       'SLA Urgence Haute (traitement) outside SLA':uho,
                       'SLA Urgence Très haute (traitement) Quantité':utq,
                       'SLA Urgence Très haute (traitement) inside SLA':uti,
                       'SLA Urgence Très haute (traitement) outside SLA':uto,
                       'Prise en charge (Support) Quantité':pecsq,
                       'Prise en charge (Support) inside SLA':pecsi,
                       'Prise en charge (Support) outside SLA':pecso,
                       'Prise en charge (Ressource) Quantité':pecdq,
                       'Prise en charge (Ressource) inside SLA':pecdi,
                       'Prise en charge (Ressource) outside SLA':pecdo,
                       'Traitement Quantité':traiq,
                       'Traitement inside SLA':traii,
                       'Traitement outside SLA':traio,
                      #Demande de la qualité en minutes
                       'Nombre de demande de support':nbinc,
                       'Nombre de demande de ressource':nbdem,
                       'Nombre d’intervention ouvert P1':nbio1,
                       'Nombre d’intervention ouvert P2':nbio2,
                       'Nombre d’intervention ouvert P3':nbio3,
                       'Nombre d’intervention ouvert P4':nbio4,
                       'Nombre de demande ouverte P1':nbdo1,
                       'Nombre de demande ouverte P2':nbdo2,
                       'Nombre de demande ouverte P3':nbdo3,
                       'Nombre de demande ouverte P4':nbdo4,
                       'Durée moyenne avant premier suivi (incident)':dmapsi,
                       'Durée moyenne avant clôture (incident)':dmaci,
           'Durée moyenne avant premier suivi (demande de ressource)':dmapsd,
           'Durée moyenne avant clôture (demande de ressource)':dmacd,
                       #Demande de Sandra
                       'Tr':tr,
                       'Tri':tri,
                       'Trd':trd,
                       'Tr1':tr1,
                       'Tr2':tr2,
                       'Tr3':tr3,
                       'Tr4':tr4,
                       'Tc1':tc1,
                       'Tc2':tc2,
                       'Tc3':tc3,
                       'Tc4':tc4
                       },ignore_index=True)
            # Typage des données avant de les inscrire en base de données,
            # cela à une influence très positive sur l’exportation dans 
            # d’autre logiciel (Power BI), qui reconnaissent automatiquement
            # le format des données grâce aux lignes ci-dessous
            da['Prise en charge (par le Front Office) Quantité']=\
            da['Prise en charge (par le Front Office) Quantité']\
            .astype(np.float64)
            da['Prise en charge (par le Front Office) inside SLA']=\
            da['Prise en charge (par le Front Office) inside SLA']\
            .astype(np.float64)
            da['Prise en charge (par le Front Office) outside SLA']=\
            da['Prise en charge (par le Front Office) outside SLA']\
            .astype(np.float64)
            da['SLA Urgence basse (traitement) Quantité']=\
            da['SLA Urgence basse (traitement) Quantité'].astype(np.int64)
            da['SLA Urgence basse (traitement) inside SLA']=\
            da['SLA Urgence basse (traitement) inside SLA'].astype(np.float64)
            da['SLA Urgence basse (traitement) outside SLA']=\
            da['SLA Urgence basse (traitement) outside SLA'].astype(np.float64)
            da['SLA Urgence moyenne (traitement) Quantité']=\
            da['SLA Urgence moyenne (traitement) Quantité'].astype(np.int64)
            da['SLA Urgence moyenne (traitement) inside SLA']=\
            da['SLA Urgence moyenne (traitement) inside SLA']\
            .astype(np.float64)
            da['SLA Urgence moyenne (traitement) outside SLA']=\
            da['SLA Urgence moyenne (traitement) outside SLA']\
            .astype(np.float64)
            da['SLA Urgence Haute (traitement) Quantité']=\
            da['SLA Urgence Haute (traitement) Quantité'].astype(np.int64)
            da['SLA Urgence Haute (traitement) inside SLA']=\
            da['SLA Urgence Haute (traitement) inside SLA'].astype(np.float64)
            da['SLA Urgence Haute (traitement) outside SLA']=\
            da['SLA Urgence Haute (traitement) outside SLA'].astype(np.float64)
            da['SLA Urgence Très haute (traitement) Quantité']=\
            da['SLA Urgence Très haute (traitement) Quantité'].astype(np.int64)
            da['SLA Urgence Très haute (traitement) inside SLA']=\
            da['SLA Urgence Très haute (traitement) inside SLA']\
            .astype(np.float64)
            da['SLA Urgence Très haute (traitement) outside SLA']=\
            da['SLA Urgence Très haute (traitement) outside SLA']\
            .astype(np.float64)
            da['Prise en charge (Support) Quantité']=\
            da['Prise en charge (Support) Quantité'].astype(np.int64)
            da['Prise en charge (Support) inside SLA']=\
            da['Prise en charge (Support) inside SLA'].astype(np.float64)
            da['Prise en charge (Support) outside SLA']=\
            da['Prise en charge (Support) outside SLA'].astype(np.float64)
            da['Prise en charge (Ressource) Quantité']=\
            da['Prise en charge (Ressource) Quantité'].astype(np.int64)
            da['Prise en charge (Ressource) inside SLA']=\
            da['Prise en charge (Ressource) inside SLA'].astype(np.float64)
            da['Prise en charge (Ressource) outside SLA']=\
            da['Prise en charge (Ressource) outside SLA'].astype(np.float64)
            da['Traitement Quantité']=\
            da['Traitement Quantité'].astype(np.int64)
            da['Traitement inside SLA']=\
            da['Traitement inside SLA'].astype(np.float64)
            da['Traitement outside SLA']=\
            da['Traitement outside SLA'].astype(np.float64)
            #Demande de la qualité en minutes
            da['Nombre de demande de support']=\
            da['Nombre de demande de support'].astype(np.int64)
            da['Nombre de demande de ressource']=\
            da['Nombre de demande de ressource'].astype(np.int64)
            da['Nombre d’intervention ouvert P1']=\
            da['Nombre d’intervention ouvert P1'].astype(np.int64)
            da['Nombre d’intervention ouvert P2']=\
            da['Nombre d’intervention ouvert P2'].astype(np.int64)
            da['Nombre d’intervention ouvert P3']=\
            da['Nombre d’intervention ouvert P3'].astype(np.int64)
            da['Nombre d’intervention ouvert P4']=\
            da['Nombre d’intervention ouvert P4'].astype(np.int64)
            da['Nombre de demande ouverte P1']=\
            da['Nombre de demande ouverte P1'].astype(np.int64)
            da['Nombre de demande ouverte P2']=\
            da['Nombre de demande ouverte P2'].astype(np.int64)
            da['Nombre de demande ouverte P3']=\
            da['Nombre de demande ouverte P3'].astype(np.int64)
            da['Nombre de demande ouverte P4']=\
            da['Nombre de demande ouverte P4'].astype(np.int64)
            da['Durée moyenne avant premier suivi (incident)']=\
            da['Durée moyenne avant premier suivi (incident)']\
            .astype(np.float64)
            da['Durée moyenne avant clôture (incident)']=\
            da['Durée moyenne avant clôture (incident)'].astype(np.float64)
            da['Durée moyenne avant premier suivi (demande de ressource)']=\
            da['Durée moyenne avant premier suivi (demande de ressource)']\
            .astype(np.float64)
            da['Durée moyenne avant clôture (demande de ressource)']=\
            da['Durée moyenne avant clôture (demande de ressource)']\
            .astype(np.float64)
            #Demande de Sandra
            da['Tr']=da['Tr'].astype(np.float64)
            da['Tri']=da['Tri'].astype(np.float64)
            da['Trd']=da['Trd'].astype(np.float64)
            da['Tr1']=da['Tr1'].astype(np.float64)
            da['Tr2']=da['Tr2'].astype(np.float64)
            da['Tr3']=da['Tr3'].astype(np.float64)
            da['Tr4']=da['Tr4'].astype(np.float64)
            da['Tc1']=da['Tc1'].astype(np.float64)
            da['Tc2']=da['Tc2'].astype(np.float64)
            da['Tc3']=da['Tc3'].astype(np.float64)
            da['Tc4']=da['Tc4'].astype(np.float64)
    # Après la boucle, sauvegarde des données en base de données et 
    # création du fichier excel contenant les données de la data frame
    # avec le suffixe du numéro de process, le process 0 est la somme de tous
    # les process, plus exactement c’est l’ensemble des tickets du mois.
    doing.to_sql(da,db="helpdesk_kpi_tables_data_kpi_process_" \
                 + str(process),mode=mode)
    da.to_excel("helpdesk_kpi_tables_data_kpi_process_" \
                + str(process) + ".xlsx")
    print('KPI s’est bien déroulé pour le process', process)
    return da

def cote_z(series):
    """
    Calcule la moyenne en heures à partir de nanosecondes, en excluant les
    données aberrantes d’une serie de la bibliothèque pandas.
    """
    z = series.mean()+ (series.std()* 2.24)
    data_purified = (pd.Series(series).loc[series < z])
    return pd.Series(data_purified).div(3600000000000).mean()

def nicky_craft():
    """
    Charge les données utiles pour établir les performances selon Nicolas E.
    puis écrit tous les mois dans un fichier excel, et sauvegarde le fichier
    avec le suffixe struct pour structure, dans le sens fond du futur fichier
    qui sera créer par la fonction excelize_nicky().
    """
    wb = load_workbook(filename = 'nicky.xlsx')
    source = wb.active
    mois = ['Janvier','Février','Mars','Avril','Mai','Juin',
                               'Juillet','Août','Septembre','Octobre','Novembre'
                               ,'Décembre']
    for var, month in enumerate(mois):
        target = wb.copy_worksheet(source)
        target.title = month
    wb.save("nicky_struct.xlsx")
    
def excelize_nicky():
    """
    Fonction pour créer le fichier excel de performance selon la vision de 
    Nicolas E. d’où le nom nicky (mnémotechnicité), avec le fichier structuré
    à cette finalité.
    """
    # Cette partie utilise la bibliothèque openpyxl, spécialisé dans la
    # création et la modélisation de fichier excel à partir d’objet Python.
    wb = load_workbook(filename = 'nicky_struct.xlsx')
    mois = ['Janvier','Février','Mars','Avril','Mai','Juin',
                               'Juillet','Août','Septembre',
                               'Octobre','Novembre'
                               ,'Décembre']
    for var, month in enumerate(mois):
        target = wb[month]
        data = pd.read_excel('liste_kpi.xlsx')
        target['D3'] = data.loc[data['mois'] == month][\
              'Prise en charge (par le Front Office) Quantité'].iloc[0]
        target['E3'] = data.loc[data['mois'] == month][\
              'Prise en charge (par le Front Office) inside SLA'].iloc[0]
        target['F3'] = data.loc[data['mois'] == month][\
              'Prise en charge (par le Front Office) outside SLA'].iloc[0]
        target['D4'] = data.loc[data['mois'] == month][\
              'SLA Urgence basse (traitement) Quantité'].iloc[0]
        target['E4'] = data.loc[data['mois'] == month][\
              'SLA Urgence basse (traitement) inside SLA'].iloc[0]
        target['F4'] = data.loc[data['mois'] == month][\
              'SLA Urgence basse (traitement) outside SLA'].iloc[0]
        target['D5'] = data.loc[data['mois'] == month][\
              'SLA Urgence moyenne (traitement) Quantité'].iloc[0]
        target['E5'] = data.loc[data['mois'] == month][\
              'SLA Urgence moyenne (traitement) inside SLA'].iloc[0]
        target['F5'] = data.loc[data['mois'] == month][\
              'SLA Urgence moyenne (traitement) outside SLA'].iloc[0]
        target['D6'] = data.loc[data['mois'] == month][\
              'SLA Urgence Haute (traitement) Quantité'].iloc[0]
        target['E6'] = data.loc[data['mois'] == month][\
              'SLA Urgence Haute (traitement) inside SLA'].iloc[0]
        target['F6'] = data.loc[data['mois'] == month][\
              'SLA Urgence Haute (traitement) outside SLA'].iloc[0]
        target['D7'] = data.loc[data['mois'] == month][\
              'SLA Urgence Très haute (traitement) Quantité'].iloc[0]
        target['E7'] = data.loc[data['mois'] == month][\
              'SLA Urgence Très haute (traitement) inside SLA'].iloc[0]
        target['F7'] = data.loc[data['mois'] == month][\
              'SLA Urgence Très haute (traitement) outside SLA'].iloc[0]
        target['D10'] = data.loc[data['mois'] == month][\
              'Prise en charge (par le Front Office) Quantité'].iloc[0]
        target['E10'] = data.loc[data['mois'] == month][\
              'Prise en charge inside SLA'].iloc[0]
        target['F10'] = data.loc[data['mois'] == month][\
              'Prise en charge outside SLA'].iloc[0]
        target['D11'] = data.loc[data['mois'] == month][\
              'Traitement Quantité'].iloc[0]
        target['E11'] = data.loc[data['mois'] == month][\
              'Traitement inside SLA'].iloc[0]
        target['F11'] = data.loc[data['mois'] == month][\
              'Traitement outside SLA'].iloc[0]
    wb.save("kpi_du_service.xlsx")
    

def extickets(db='helpdesk_data_mars'):
    """
    C’est la fonction qui crée la source de données premières à partir de
    tous les objets informatiques de la classe Ticket. C’est la source de
    données qui est construit sur la base de données du service. Toute
    modification de la structure de la base de données de production du SI 
    pourrait avoir des répercutions sur l’exécution de cette fonction. Cette
    fonction est en amont de tous les calculs de performances qui suivrons. Le
    nom de la table et du fichier excel qui contiennent ses valeurs porte le 
    suffixe data car c’est la source première de données qui sert à tous les 
    autres calculs.
    """
    # Création préalable de la dataframe, ne comporte pas toutes les colonnes
    # car elles seront ajoutés dans la boucle. Cependant, la dataframe doit
    # impérativement être déclarée avant la boucle pour éviter le problème
    # d’espace de nom ou "scope" et ainsi utiliser le même objet df hors et 
    # dedans la boucle for.
    df = pd.DataFrame({
            'demande': [], 
            'urgence': [],
            'creation': [],
            '1ere_pec':[],
            'type':[],
            #'Attente Niv 1': [],
            'str_type':[],
            })


    for i in range(1,doing.g_row_nb()):
        ticket = doing.instantiate_a_ticket(i)
        if ticket is not None:
            #calcul du temps de traitement approprié pour pallier l’horloge hors heure de bureau
            if (ticket.Tt.value == 0)&(ticket.a_une_cloture == True)&(ticket.est_resolu_en_suivi == True):
                    ticket.Tt = ticket.tr + ticket.tc + ticket.t1 + ticket.t2 + ticket.t3
            if (ticket.Tt.value == 0)&(ticket.a_une_cloture == True)&(ticket.est_resolu_en_suivi == False):
                    ticket.Tt = ticket.trc + ticket.t1 + ticket.t2 + ticket.t3
            df = df.append({\
                            'demande': ticket.demande, \
                            'urgence': ticket.urgence, \
                            'groupe': ticket.groupe,\
                            'str_groupe': str(trn.nomgrp(ticket.groupe)),
                            'clos': ticket.a_une_cloture,\
                            'resolu':ticket.est_resolu_en_suivi,\
                            'creation':ticket.creation,\
                            '1ere_pec': ticket.liste_date_pec[0] \
                            if ticket.a_une_pec else None,\
                            'date_de_cloture': ticket.date_de_cloture \
                            if ticket.a_une_cloture else None,\
                            'str_type': str(trn.nomticket(ticket.type)),
                            'type':ticket.type,
                            #'traitement': ticket.Tt,
                            'T1': ticket.t1.value if ticket.a_une_pec \
                            else None,
                            'T2': ticket.t2.value if ticket.a_une_pec \
                            else None,
                            'T3': ticket.t3.value if ticket.a_une_pec \
                            else None,
                            'Tc': ticket.tc.value,
                            'Tr':ticket.tr.value,
                            'Trc':ticket.trc.value,
                            'Tt':ticket.Tt.value,
                            'Process':ticket.str_process,
                            'Tech_clos':int(ticket.tech_de_cloture),
                            'Tech_res':int(ticket.tech_de_res)
                            }, ignore_index=True)
    # Formatage des tonnées suivant leur type, améliore la compatibilité
    # avec d’autres programmes utilisant ces mêmes données.
    df['demande'] = df['demande'].astype(int)
    df['urgence'] = df['urgence'].astype(int)
    df['groupe'] = df['groupe'].astype(int)
    df['str_groupe'] = df['str_groupe'].astype(str)
    df['clos'] = df['clos'].astype(bool)
    df['resolu'] = df['resolu'].astype(bool)
    df['creation'] = df['creation'].astype(str)
    df['date_de_cloture'] = df['date_de_cloture'].astype(str)
    df['1ere_pec'] = df['1ere_pec'].astype(str)
    #df['derniere_resolution'] = df['derniere_resolution'].astype(str)
    df['str_type'] = df['str_type'].astype(str)
    df['type'] = df['type'].astype(int)
    df['T1'] = df['T1'].astype(float)
    df['T2'] = df['T2'].astype(float)
    df['T3'] = df['T3'].astype(float)
    df['Tc'] = df['Tc'].astype(float)
    df['Tr'] = df['Tr'].astype(float)
    df['Trc'] = df['Trc'].astype(float)
    df['Tt'] = df['Tt'].astype(float)
    df['Process'] = df['Process'].astype(str) 
    df['Tech_clos'] = df['Tech_clos'].astype(str) 
    df['Tech_res'] = df['Tech_res'].astype(str) 

    df.to_excel('data.xlsx')

    df.index = df['demande']
    del df['demande']
    doing.to_sql(df, db="helpdesk_kpi_tables_data")
    print('Tickets expression s’est bien déroulé')
    return df