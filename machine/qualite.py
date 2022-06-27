# -*- coding: utf-8 -*-
"""
Created on Mon Mar 16 14:25:28 2020

@author: rboyrie
"""

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension

import pandas as pd
import numpy as np
from machine import do

def make():
    qualite_craft()
    excelize_qualite()
    print('Fichier créé de la qualité')

def qualite_craft():
    wb = load_workbook(filename = 'structure_7.xlsx')
    source = wb.active
    #mois = ['Janvier','Février','Mars','Avril','Mai','Juin',
    #                           'Juillet','Août','Septembre','Octobre','Novembre'
    #                           ,'Décembre']
    # Noter les mois que l’on va observer
    mois = np.array([['2019-10-01','OCT 19'],
                      ['2019-11-01','NOV 19'],
                      ['2019-12-01','DEC 19'],
                      ['2020-01-01', 'JANV 20'],
                      ['2020-02-01','FEV 20'],
                      ['2020-03-01','MAR 20'],
                      ['2020-04-01','AVR 20'],
                      ['2020-05-01','MAI 20'],
                      ['2020-06-01','JUIN 20']])
    
   

   # for var, month in enumerate(mois):
   #     target = wb.copy_worksheet(source)
    #    target.title = month[1]
    wb.save("structure_mensualise.xlsx")

def excelize_qualite():
    wb = load_workbook(filename = 'structure_mensualise.xlsx')
    #source = wb.active
    #target = wb.copy_worksheet(source)
    #target.title = month
#    mois = ['Janvier','Février','Mars','Avril','Mai','Juin',
 #                              'Juillet','Août','Septembre','Octobre','Novembre'
  #                             ,'Décembre']
    #mois = ['2019/01/01','2019/02/01','2019/03/01','2019/04/01','2019/05/01','2019/06/01','2019/06/01','2019/07/01','2019/08/01','2019/10/01','2019/11/01','2019/12/01','2020/01/01','2020/02/01']
    # Noter les mois que l’on va observer
    mois = np.array([['2019-10-01','OCT 19','D'],
                      ['2019-11-01','NOV 19','E'],
                      ['2019-12-01','DEC 19','F'],
                      ['2020-01-01', 'JANV 20','G'],
                      ['2020-02-01','FEV 20','H'],
                      ['2020-03-01','MAR 20','I'],
                      ['2020-04-01','AVR 20','J'],
                      ['2020-05-01','MAI 20','K'],
                      ['2020-06-01','JUIN 20','L']])
   
    for var, month in enumerate(mois):
        target = wb[month[1]]
        synthese = wb['SYNTHESE 20']
        data1 = pd.read_excel('helpdesk_kpi_tables_data_kpi_process_1.xlsx')
        data2 = pd.read_excel('helpdesk_kpi_tables_data_kpi_process_2.xlsx')
        data3 = pd.read_excel('helpdesk_kpi_tables_data_kpi_process_3.xlsx')
       # data4 = pd.read_excel('helpdesk_kpi_tables_data_kpi_process_4.xlsx')
        if var > 0:
           target['D8'] = str(mois[var-1][1])
        target['E8'] = str(month[1])
        target['E14'] = data1.loc[data1['date'] == month[0]]['Nombre de demande de support'].iloc[0]
        target['E15'] = data1.loc[data1['date'] == month[0]]['Nombre d’intervention ouvert P1'].iloc[0]
        target['E16'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
        target['E17'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant clôture (incident)'].iloc[0]
        #target['E18'] = float(int(target['E14'].internal_value)/int(target['E15'].internal_value))
        target['E19'] = data1.loc[data1['date'] == month[0]]['Nombre de demande de ressource'].iloc[0]
        target['E20'] = data1.loc[data1['date'] == month[0]]['Nombre de demande ouverte P1'].iloc[0]
        target['E21'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
        target['E22'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
        #target['E23'] = 1 if (int(target['E20'].internal_value)>int(target['E19'].internal_value)) else float(int(target['E20'].internal_value)/int(target['E19'].internal_value))
        synthese[month[2]+'8'] = str(month[1])
        synthese[month[2]+'14'] = data1.loc[data1['date'] == month[0]]['Nombre de demande de support'].iloc[0]
        synthese[month[2]+'15'] = data1.loc[data1['date'] == month[0]]['Nombre d’intervention ouvert P1'].iloc[0]
        synthese[month[2]+'16'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
        synthese[month[2]+'17'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant clôture (incident)'].iloc[0]
        #synthese[month[2]+'18'] = float(int(synthese[month[2]+'14'].internal_value)/int(synthese[month[2]+'15'].internal_value))
        synthese[month[2]+'19'] = data1.loc[data1['date'] == month[0]]['Nombre de demande de ressource'].iloc[0]
        synthese[month[2]+'20'] = data1.loc[data1['date'] == month[0]]['Nombre de demande ouverte P1'].iloc[0]
        synthese[month[2]+'21'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
        synthese[month[2]+'22'] = data1.loc[data1['date'] == month[0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
        #synthese[month[2]+'23'] = 1 if (int(synthese[month[2]+'20'].internal_value)>int(synthese[month[2]+'19'].internal_value)) else float(int(synthese[month[2]+'20'].internal_value)/int(synthese[month[2]+'19'].internal_value))
        if var > 0:
            target['D14'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre de demande de support'].iloc[0]
            target['D15'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre d’intervention ouvert P1'].iloc[0]
            target['D16'] = data1.loc[data1['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
            target['D17'] = data1.loc[data1['date'] == mois[var-1][0]]['Durée moyenne avant clôture (incident)'].iloc[0]
            target['D18'] = float(int(target['D14'].internal_value)/int(target['D15'].internal_value))
            target['D19'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre de demande de ressource'].iloc[0]
            target['D20'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre de demande ouverte P1'].iloc[0]
            target['D21'] = data1.loc[data1['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
            target['D22'] = data1.loc[data1['date'] == mois[var-1][0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
         #   target['D23'] = 1 if (int(target['D20'].internal_value)>int(target['D19'].internal_value)) else float(int(target['D20'].internal_value)/int(target['D19'].internal_value))    
        target['E25'] = data2.loc[data2['date'] == month[0]]['Nombre de demande de support'].iloc[0]
        target['E26'] = data2.loc[data2['date'] == month[0]]['Nombre d’intervention ouvert P2'].iloc[0]
        target['E27'] = data2.loc[data2['date'] == month[0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
        target['E28'] = data2.loc[data2['date'] == month[0]]['Durée moyenne avant clôture (incident)'].iloc[0]
        target['E30'] = data2.loc[data2['date'] == month[0]]['Nombre de demande de ressource'].iloc[0]
        target['E31'] = data1.loc[data1['date'] == month[0]]['Nombre de demande ouverte P2'].iloc[0]
        target['E32'] = data2.loc[data2['date'] == month[0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
        target['E33'] = data2.loc[data2['date'] == month[0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
        if var > 0:
            target['D25'] = data2.loc[data2['date'] == mois[var-1][0]]['Nombre de demande de support'].iloc[0]
            target['D26'] = data2.loc[data2['date'] == mois[var-1][0]]['Nombre d’intervention ouvert P2'].iloc[0]
            target['D27'] = data2.loc[data2['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
            target['D28'] = data2.loc[data2['date'] == mois[var-1][0]]['Durée moyenne avant clôture (incident)'].iloc[0]
            target['D30'] = data2.loc[data2['date'] == mois[var-1][0]]['Nombre de demande de ressource'].iloc[0]
            target['D31'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre de demande ouverte P2'].iloc[0]  
            target['D32'] = data2.loc[data2['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
            target['D33'] = data2.loc[data2['date'] == mois[var-1][0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
                
        target['E36'] = data3.loc[data3['date'] == month[0]]['Nombre de demande de support'].iloc[0]
        target['E37'] = data3.loc[data3['date'] == month[0]]['Nombre d’intervention ouvert P3'].iloc[0]
        target['E38'] = data3.loc[data3['date'] == month[0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
        target['E39'] = data3.loc[data3['date'] == month[0]]['Durée moyenne avant clôture (incident)'].iloc[0]
        target['E41'] = data3.loc[data3['date'] == month[0]]['Nombre de demande de ressource'].iloc[0]
        target['E42'] = data1.loc[data1['date'] == month[0]]['Nombre de demande ouverte P3'].iloc[0]  
        target['E43'] = data3.loc[data3['date'] == month[0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
        target['E44'] = data3.loc[data3['date'] == month[0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
        if var > 0:
            target['D36'] = data3.loc[data3['date'] == mois[var-1][0]]['Nombre de demande de support'].iloc[0]
            target['D37'] = data3.loc[data3['date'] == mois[var-1][0]]['Nombre d’intervention ouvert P3'].iloc[0]
            target['D38'] = data3.loc[data3['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
            target['D39'] = data3.loc[data3['date'] == mois[var-1][0]]['Durée moyenne avant clôture (incident)'].iloc[0]
            target['D41'] = data3.loc[data3['date'] == mois[var-1][0]]['Nombre de demande de ressource'].iloc[0]
            target['D42'] = data1.loc[data1['date'] == mois[var-1][0]]['Nombre de demande ouverte P3'].iloc[0]  
            target['D43'] = data3.loc[data3['date'] == mois[var-1][0]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
            target['D444'] = data3.loc[data3['date'] == mois[var-1][0]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
        
#        target['E35'] = data4.loc[data4['date'] == month]['Nombre de demande de support'].iloc[0]
#        target['E36'] = data4.loc[data4['date'] == month]['Nombre d’intervention ouvert P4'].iloc[0]
#        target['E37'] = data4.loc[data4['date'] == month]['Durée moyenne avant premier suivi (incident)'].iloc[0]
#        target['E38'] = data4.loc[data4['date'] == month]['Durée moyenne avant clôture (incident)'].iloc[0]
#        target['E39'] = data4.loc[data4['date'] == month]['Nombre de demande de ressource'].iloc[0]
#        target['E40'] = data4.loc[data4['date'] == month]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
#        target['E41'] = data4.loc[data4['date'] == month]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
#        if var > 0:
#            target['D35'] = data4.loc[data4['date'] == mois[var-1]]['Nombre de demande de support'].iloc[0]
#            target['D36'] = data4.loc[data4['date'] == mois[var-1]]['Nombre d’intervention ouvert P4'].iloc[0]
#            target['D37'] = data4.loc[data4['date'] == mois[var-1]]['Durée moyenne avant premier suivi (incident)'].iloc[0]
#            target['D38'] = data4.loc[data4['date'] == mois[var-1]]['Durée moyenne avant clôture (incident)'].iloc[0]
#            target['D39'] = data4.loc[data4['date'] == mois[var-1]]['Nombre de demande de ressource'].iloc[0]
#            target['D40'] = data4.loc[data4['date'] == mois[var-1]]['Durée moyenne avant premier suivi (demande de ressource)'].iloc[0]
#            target['D41'] = data4.loc[data4['date'] == mois[var-1]]['Durée moyenne avant clôture (demande de ressource)'].iloc[0]
    wb.save("Qualité_SI_new.xlsx")
    
def school():
    """
    Bulletin scolaire de résultat technicien et client
    """
    wb = Workbook()
    source = do.r_to_b(query='SELECT * FROM supply_demand_par_tech_proportion')
    source.index = source['date']
    source = source.drop('date',axis=1)
    #print(source)
    source = source.rename(columns={'qut':'Quantité de tickets',
                                               'qb':'Urgence Basse','qm':'Urgence Moyenne','qh':'Urgence Haute','qt':'Urgence Très Haute',
                                               'qrbi':'Résolution In urgence basse','qrmi':'Résolution In urgence moyenne',
                                               'qrhi':'Résolution In urgence haute','qrti':'Résolution In urgence très haute',
                                               'qtbi':'Traitement In urgence basse',
                                               'qtmi':'Traitement In urgence moyenne',
                                               'qthi':'Traitement In urgence haute',
                                               'qtti':'Traitement In urgence très haute',
                                               'mean_resolution':'Temps moyen de résolution',
                                               'mean_traitement':'Temps moyen de traitement','process':'process', 'str_type': 'type', 'tech':'nom'})
    liste = source['nom'].unique()
    for i,tech in enumerate(liste):
        #print (i , tech)
        value = source.loc[(source['nom'].str.contains(tech))]
        #print(value)
        ws = wb.create_sheet(title=tech)
        rows = dataframe_to_rows(value)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)
        ws.column_dimensions['A'].width=20
        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20
        ws.column_dimensions['F'].width=20
        ws.column_dimensions['G'].width=20
        ws.column_dimensions['H'].width=20
        ws.column_dimensions['I'].width=20
        ws.column_dimensions['J'].width=20
        ws.column_dimensions['K'].width=20
        ws.column_dimensions['L'].width=20
        ws.column_dimensions['M'].width=20
        ws.column_dimensions['N'].width=20
    wb._sheets.sort(key=lambda ws: ws.title)
    del wb['Sheet']

    wb.save("Bulletin_technicien.xlsx")

    #sheet = wb.get_sheet_by_name('spam')
    
    wb = Workbook()
    source = do.r_to_b(query='SELECT * FROM supply_demand_par_client_proportion')
    source.index = source['date']
    source = source.drop('date',axis=1)
    source= source.rename(columns={'qut':'Quantité de tickets',
                                               'qb':'Urgence Basse','qm':'Urgence Moyenne','qh':'Urgence Haute','qt':'Urgence Très Haute',
                                               'qrbi':'Résolution In urgence basse','qrmi':'Résolution In urgence moyenne',
                                               'qrhi':'Résolution In urgence haute','qrti':'Résolution In urgence très haute',
                                               'qtbi':'Traitement In urgence basse',
                                               'qtmi':'Traitement In urgence moyenne',
                                               'qthi':'Traitement In urgence haute',
                                               'qtti':'Traitement In urgence très haute',
                                               'mean_resolution':'Temps moyen de résolution',
                                               'mean_traitement':'Temps moyen de traitement','process':'process', 'str_type': 'type', 'client':'nom'})
    liste = source['nom'].unique()
    
    for i,tech in enumerate(liste):
        #print (i , tech)
        value = source.loc[(source['nom'].str.contains(tech))]
        #print(value)
        ws = wb.create_sheet(title=tech)
        rows = dataframe_to_rows(value)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)
                 
        ws.column_dimensions['A'].width=20
        ws.column_dimensions['B'].width=20
        ws.column_dimensions['C'].width=20
        ws.column_dimensions['D'].width=20
        ws.column_dimensions['E'].width=20
        ws.column_dimensions['F'].width=20
        ws.column_dimensions['G'].width=20
        ws.column_dimensions['H'].width=20
        ws.column_dimensions['I'].width=20
        ws.column_dimensions['J'].width=20
        ws.column_dimensions['K'].width=20
        ws.column_dimensions['L'].width=20
        ws.column_dimensions['M'].width=20
        ws.column_dimensions['N'].width=20
    wb._sheets.sort(key=lambda ws: ws.title)
    del wb['Sheet']
    wb.save("Bulletin_client.xlsx")